from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import datetime
import openpyxl
import os
import sys
import time

import pandas as pd

# =============================================================================================== #
# BASE UTILS                                                                                      #
# =============================================================================================== #

def get_workbook() -> Workbook:
    return Workbook()

def save_workbook(workbook: Workbook, filename: str) -> None:
    workbook.save(f"{filename}")
    pass

def close_workbook(workbook: Workbook) -> None:
    workbook.close()
    pass

def create_sheet(workbook: Workbook, sheet_name: str, index: int = None) -> object: 
    return workbook.create_sheet(f'{sheet_name}', index=index)

def get_sheet_by_name(workbook: Workbook, sheet_name: str) -> object: 
    return workbook[f'{sheet_name}']

# =============================================================================================== #
# ADD HISTORY TO SPREDSHEET                                                                       #
# =============================================================================================== #

def add_history_2_spredsheet(path_2_history: str, columns: list, workbook: Workbook, sheet_name: str = 'History') -> None:
    sheet_history = create_sheet(workbook, sheet_name, index=1)

    df_history: pd.DataFrame = pd.read_csv(path_2_history, skiprows=1, names=columns)
    result = df_history.describe()

    for r in dataframe_to_rows(result, index=True, header=True):
        sheet_history.append(r)
        pass
    
    sheet_history.append([])
    sheet_history.append([])
    
    for r in dataframe_to_rows(df_history, index=True, header=True):
        sheet_history.append(r)
        pass

    pass

# =============================================================================================== #
# ADD INPUT PARAMS TO SPREDSHEET                                                                  #
# =============================================================================================== #

def _append_keys_and_values_2_sheet(sheet, tmp_dict: dict, col: int, header: list) -> None:
    # keys: list = list(tmp_dict.keys())
    # values: list = list(map(lambda xi: str(xi), tmp_dict.values()))

    # sheet.append(keys)
    # sheet.append(values)
    start_offset: int = 2
    for ii, xx in enumerate(header):
        sheet.cell(row=start_offset, column=col+ii).value = xx
    start_offset += 1
    for ii, (k, v) in enumerate(tmp_dict.items()):
        sheet.cell(row=start_offset+ii, column=col).value = str(k)
        sheet.cell(row=start_offset+ii, column=col+1).value = str(v)
    pass

def add_input_params_2_spredsheet(cmd_args_dict: dict, network_input_params_dict: dict, workbook: Workbook, sheet_name: str = 'Input Params') -> None:
    sheet_input_params = create_sheet(workbook, sheet_name, index=0)

    _append_keys_and_values_2_sheet(sheet_input_params, cmd_args_dict, col=2, header=['Name Input Params', 'Value'])
    _append_keys_and_values_2_sheet(sheet_input_params, network_input_params_dict, col=7, header=['Name NN Params', 'Value'])
    pass

# =============================================================================================== #
# ADD IMAGES TO SPREDSHEET                                                                        #
# =============================================================================================== #

def add_image_2_spredsheet(image_path: str, workbook: Workbook, sheet_name: str = 'Neural Network Graph', anchor: str = 'A1'):
    sheet_graph_nn = create_sheet(workbook, sheet_name, index=2)

    img = openpyxl.drawing.image.Image(image_path)
    img.anchor = anchor
    sheet_graph_nn.add_image(img)
    pass

# ------------------------------- #
# Links                           #
# ------------------------------- #

# https://openpyxl.readthedocs.io/en/stable/index.html
# https://realpython.com/openpyxl-excel-spreadsheets-python/